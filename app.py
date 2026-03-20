import io
import math
import os
import re
import zipfile
from copy import copy
from datetime import datetime
from tempfile import TemporaryDirectory

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

st.set_page_config(page_title="补货单生成工具", layout="wide")

YELLOW_FILL = PatternFill(fill_type="solid", fgColor="FFF59D")
ERROR_FILL = PatternFill(fill_type="solid", fgColor="FDE2E1")


def safe_set(ws, row, col, value):
    cell = ws.cell(row, col)
    if isinstance(cell, MergedCell):
        return False
    cell.value = value
    return True


def clear_non_merged_range(ws, start_row, end_row, start_col, end_col):
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(r, c)
            if isinstance(cell, MergedCell):
                continue
            cell.value = None
            cell.fill = copy(cell.fill)


def normalize_text(v):
    if pd.isna(v) or v is None:
        return ""
    s = str(v).strip()
    s = s.replace("（", "(").replace("）", ")")
    s = s.replace("【", "(").replace("】", ")")
    s = s.replace(" ", " ").replace("\u3000", " ")
    s = re.sub(r"\s+", "", s)
    return s.upper()


def normalize_model(v):
    s = normalize_text(v)
    s = s.replace("OPPO", "").replace("ONEPLUS", "")
    s = s.replace("5G", "5G")
    return s


def normalize_color(v):
    return normalize_text(v)


def extract_purchase_order_no(order_no):
    if order_no is None:
        return ""
    s = str(order_no)
    m = re.search(r"(\d{10})", s)
    return m.group(1) if m else s


def region_from_remark(remark):
    s = str(remark or "")
    if "广州" in s:
        return "广州分销"
    if "粤北" in s:
        return "粤北分销"
    return None


def build_aux_lookup(df_aux):
    lookup = {}
    for _, row in df_aux.iterrows():
        model = normalize_model(row.get("CTMS机型"))
        color = normalize_color(row.get("颜色"))
        if model and color:
            lookup[(model, color)] = {
                "code": "" if pd.isna(row.get("物料编码")) else str(row.get("物料编码")).strip(),
                "desc": "" if pd.isna(row.get("SCM物料描述")) else str(row.get("SCM物料描述")).strip(),
            }
    return lookup


def match_aux(row, lookup):
    model = normalize_model(row.get("型号"))
    color = normalize_color(row.get("颜色"))
    exact = lookup.get((model, color))
    if exact:
        return exact, ""

    # 兜底：型号包含关系匹配
    candidates = []
    for (m, c), data in lookup.items():
        if c == color and (model in m or m in model):
            candidates.append((m, data))
    if len(candidates) == 1:
        return candidates[0][1], "模糊型号匹配"
    if len(candidates) > 1:
        return {"code": "", "desc": ""}, "辅助表存在多个候选"
    return {"code": "", "desc": ""}, "辅助表未匹配"


def prepare_region_df(df_source, region_name, aux_lookup):
    df = df_source.copy()
    df["区域"] = df["备注"].apply(region_from_remark)
    df = df[df["区域"] == region_name].copy()
    df = df[df["本次配送台数"].fillna(0) != 0].copy()

    rows = []
    missing = []
    for idx, row in df.reset_index(drop=True).iterrows():
        aux, note = match_aux(row, aux_lookup)
        code = aux.get("code", "")
        desc = aux.get("desc", "")
        if not code or not desc:
            missing.append({
                "序号": idx + 1,
                "型号": row.get("型号", ""),
                "颜色": row.get("颜色", ""),
                "问题": note or "缺少物料编码/描述",
            })
        rows.append({
            "#": idx + 1,
            "终端公司采购单号": extract_purchase_order_no(row.get("补货单号")),
            "品牌": row.get("品牌", ""),
            "编码": code,
            "产品描述": desc,
            "配送单总台数": int(row.get("配送单总台数", 0) or 0),
            "本次配送台数": int(row.get("本次配送台数", 0) or 0),
            "箱数": math.ceil((float(row.get("本次配送台数", 0) or 0)) / 10) if row.get("本次配送台数", 0) else 0,
            "补货单号": row.get("补货单号", ""),
            "指令编号": "",
            "备注": row.get("备注", ""),
            "_missing": bool(not code or not desc),
        })
    return pd.DataFrame(rows), pd.DataFrame(missing)


def fill_template_sheet(ws, region_df, missing_df):
    # 清空明细区与合计区（跳过合并单元格）
    clear_non_merged_range(ws, 8, 20, 1, 11)
    clear_non_merged_range(ws, 21, 24, 1, 11)

    today = datetime.today()
    safe_set(ws, 2, 3, today)

    start_row = 8
    for i, (_, row) in enumerate(region_df.iterrows(), start=0):
        excel_row = start_row + i
        data = [
            row["#"], row["终端公司采购单号"], row["品牌"], row["编码"], row["产品描述"],
            row["配送单总台数"], row["本次配送台数"], row["箱数"], row["补货单号"], row["指令编号"], row["备注"]
        ]
        for col_idx, value in enumerate(data, start=1):
            ok = safe_set(ws, excel_row, col_idx, value)
            if ok and col_idx in (4, 5) and row["_missing"]:
                ws.cell(excel_row, col_idx).fill = copy(YELLOW_FILL)

    total_row = 21
    safe_set(ws, total_row, 1, "合计：")
    safe_set(ws, total_row, 6, int(region_df["配送单总台数"].sum()) if not region_df.empty else 0)
    safe_set(ws, total_row, 7, int(region_df["本次配送台数"].sum()) if not region_df.empty else 0)
    safe_set(ws, total_row, 8, int(region_df["箱数"].sum()) if not region_df.empty else 0)
    safe_set(ws, 22, 1, "签收人：")
    safe_set(ws, 23, 1, "签收数量：")
    safe_set(ws, 24, 1, "签收日期、时间：")

    # 超出模板 13 行时给提示（当前模板最多支持 13 条）
    if len(region_df) > 13:
        safe_set(ws, 24, 6, f"提示：当前共{len(region_df)}条，超过模板13条，建议扩展模板")
        ws.cell(24, 6).fill = copy(ERROR_FILL)


def build_pdf(pdf_path, region_name, region_df, missing_df):
    pdfmetrics.registerFont(UnicodeCIDFont('STSong-Light'))
    doc = SimpleDocTemplate(
        pdf_path,
        pagesize=landscape(A4),
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
    )
    styles = getSampleStyleSheet()
    title_style = styles["Title"]
    title_style.fontName = 'STSong-Light'
    title_style.fontSize = 16
    normal = styles["Normal"]
    normal.fontName = 'STSong-Light'
    normal.fontSize = 9
    normal.leading = 12

    story = []
    story.append(Paragraph("厂家物流送货清单打印", title_style))
    story.append(Spacer(1, 4))
    info_lines = [
        f"预计送货日期：{datetime.today().strftime('%Y-%m-%d')}",
        "发货单位：重庆康久盛通讯设备有限公司    收货单位：中国移动通信集团终端有限公司广东分公司",
        "发货仓地址：广东省广州市花都区花东镇金谷工业园永大路7号10平台",
        "收货仓地址：广东省广州市花都区花东镇金港北一路3号J8栋301单元",
        f"单据区域：{region_name}",
    ]
    for line in info_lines:
        story.append(Paragraph(line, normal))
    story.append(Spacer(1, 6))

    headers = ["#", "终端公司采购单号", "品牌", "编码", "产品描述", "配送单总台数", "本次配送台数", "箱数", "补货单号", "备注"]
    table_data = [headers]
    for _, row in region_df.iterrows():
        table_data.append([
            row["#"], row["终端公司采购单号"], row["品牌"], row["编码"], row["产品描述"],
            row["配送单总台数"], row["本次配送台数"], row["箱数"], row["补货单号"], row["备注"]
        ])
    table_data.append(["合计", "", "", "", "", int(region_df["配送单总台数"].sum()) if not region_df.empty else 0,
                       int(region_df["本次配送台数"].sum()) if not region_df.empty else 0,
                       int(region_df["箱数"].sum()) if not region_df.empty else 0, "", ""])

    col_widths = [12*mm, 30*mm, 16*mm, 34*mm, 80*mm, 22*mm, 22*mm, 14*mm, 38*mm, 24*mm]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    style = TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), 'STSong-Light'),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("LEADING", (0, 0), (-1, -1), 10),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EAEFF7")),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#F7F7F7")),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("ALIGN", (1, 0), (3, -1), "CENTER"),
        ("ALIGN", (5, 0), (7, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ])
    # 标记缺失
    for i, (_, row) in enumerate(region_df.iterrows(), start=1):
        if row.get("_missing"):
            style.add("BACKGROUND", (3, i), (4, i), colors.HexColor("#FFF59D"))
    table.setStyle(style)
    story.append(table)
    story.append(Spacer(1, 8))
    story.append(Paragraph("签收人：________________    签收数量：________________    签收日期、时间：________________", normal))

    if not missing_df.empty:
        story.append(Spacer(1, 8))
        story.append(Paragraph("需复核的辅助表匹配：", normal))
        miss_data = [["序号", "型号", "颜色", "问题"]] + missing_df.astype(str).values.tolist()
        miss_table = Table(miss_data, colWidths=[15*mm, 70*mm, 30*mm, 50*mm], repeatRows=1)
        miss_table.setStyle(TableStyle([
            ("FONTNAME", (0,0), (-1,-1), 'STSong-Light'),
            ("FONTSIZE", (0,0), (-1,-1), 8),
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#FDE2E1")),
            ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ]))
        story.append(miss_table)

    doc.build(story)


def process_file(uploaded_file):
    file_bytes = uploaded_file.getvalue()
    df_source = pd.read_excel(io.BytesIO(file_bytes), sheet_name="源文件", header=1)
    df_aux = pd.read_excel(io.BytesIO(file_bytes), sheet_name="辅助表")

    aux_lookup = build_aux_lookup(df_aux)
    results = {}

    with TemporaryDirectory() as tmpdir:
        for region in ["广州分销", "粤北分销"]:
            region_df, missing_df = prepare_region_df(df_source, region, aux_lookup)
            out_wb = load_workbook(io.BytesIO(file_bytes))
            ws = out_wb["生成pdf模板"]
            fill_template_sheet(ws, region_df, missing_df)

            xlsx_name = f"{region}补货单.xlsx"
            pdf_name = f"{region}补货单.pdf"
            xlsx_path = os.path.join(tmpdir, xlsx_name)
            pdf_path = os.path.join(tmpdir, pdf_name)
            out_wb.save(xlsx_path)
            build_pdf(pdf_path, region, region_df, missing_df)

            with open(xlsx_path, "rb") as f:
                xlsx_bytes = f.read()
            with open(pdf_path, "rb") as f:
                pdf_bytes = f.read()

            results[region] = {
                "rows": len(region_df),
                "missing": len(missing_df),
                "missing_df": missing_df,
                "xlsx_name": xlsx_name,
                "xlsx_bytes": xlsx_bytes,
                "pdf_name": pdf_name,
                "pdf_bytes": pdf_bytes,
            }

    return results


def build_zip(results):
    mem = io.BytesIO()
    with zipfile.ZipFile(mem, "w", zipfile.ZIP_DEFLATED) as zf:
        for _, item in results.items():
            zf.writestr(item["xlsx_name"], item["xlsx_bytes"])
            zf.writestr(item["pdf_name"], item["pdf_bytes"])
    mem.seek(0)
    return mem.getvalue()


st.title("OPPO 补货单自动生成工具")
st.caption("上传包含【源文件】【生成pdf模板】【辅助表】三个 sheet 的 Excel，自动生成广州分销 / 粤北分销两套 Excel 和 PDF。")

uploaded_file = st.file_uploader("上传源文件", type=["xlsx"])

if uploaded_file:
    if st.button("一键生成", type="primary"):
        try:
            results = process_file(uploaded_file)
            summary = []
            for region, item in results.items():
                missing_preview = "无"
                if item["missing"]:
                    missing_preview = "；".join(
                        f"{r['型号']} / {r['颜色']} / {r['问题']}" for _, r in item["missing_df"].head(5).iterrows()
                    )
                summary.append({
                    "区域": region,
                    "明细行数": item["rows"],
                    "缺失匹配数": item["missing"],
                    "缺失匹配明细": missing_preview,
                })

            st.subheader("生成结果")
            st.dataframe(pd.DataFrame(summary), use_container_width=True)

            col1, col2 = st.columns(2)
            with col1:
                gz = results["广州分销"]
                st.markdown("### 广州分销")
                st.download_button("下载广州 Excel", gz["xlsx_bytes"], file_name=gz["xlsx_name"], mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                st.download_button("下载广州 PDF", gz["pdf_bytes"], file_name=gz["pdf_name"], mime="application/pdf")
            with col2:
                yb = results["粤北分销"]
                st.markdown("### 粤北分销")
                st.download_button("下载粤北 Excel", yb["xlsx_bytes"], file_name=yb["xlsx_name"], mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                st.download_button("下载粤北 PDF", yb["pdf_bytes"], file_name=yb["pdf_name"], mime="application/pdf")

            zip_bytes = build_zip(results)
            st.download_button("一键下载全部（ZIP）", zip_bytes, file_name="补货单生成结果.zip", mime="application/zip")

            for region, item in results.items():
                if item["missing"]:
                    with st.expander(f"{region} - 未匹配辅助表明细"):
                        st.dataframe(item["missing_df"], use_container_width=True)
        except Exception as e:
            st.error(f"生成失败：{e}")
